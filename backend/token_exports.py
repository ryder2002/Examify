"""Encrypted activation-code storage and XLSX exports for administrators."""

from __future__ import annotations

import base64
import hashlib
import os
from io import BytesIO
from typing import Any, Iterable

from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class TokenExportUnavailable(RuntimeError):
    pass


def _key(secret: str) -> bytes:
    if len(secret) < 32:
        raise TokenExportUnavailable(
            "TOKEN_EXPORT_SECRET phải có ít nhất 32 ký tự"
        )
    return hashlib.sha256(secret.encode("utf-8")).digest()


def encrypt_activation_code(code: str, token_id: str, secret: str) -> str:
    nonce = os.urandom(12)
    ciphertext = AESGCM(_key(secret)).encrypt(
        nonce, code.encode("utf-8"), token_id.encode("utf-8")
    )
    return "v1.{}.{}".format(
        base64.urlsafe_b64encode(nonce).decode("ascii").rstrip("="),
        base64.urlsafe_b64encode(ciphertext).decode("ascii").rstrip("="),
    )


def decrypt_activation_code(value: str, token_id: str, secret: str) -> str:
    try:
        version, raw_nonce, raw_ciphertext = value.split(".", 2)
        if version != "v1":
            raise ValueError("unsupported version")

        def decode(item: str) -> bytes:
            return base64.urlsafe_b64decode(item + "=" * (-len(item) % 4))

        plaintext = AESGCM(_key(secret)).decrypt(
            decode(raw_nonce),
            decode(raw_ciphertext),
            token_id.encode("utf-8"),
        )
        return plaintext.decode("utf-8")
    except TokenExportUnavailable:
        raise
    except Exception as exc:
        raise TokenExportUnavailable("Không giải mã được token") from exc


def _excel_safe(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def build_token_workbook(rows: Iterable[dict[str, Any]]) -> bytes:
    all_rows = list(rows)
    workbook = Workbook()
    workbook.remove(workbook.active)
    columns = [
        ("STT", "index"),
        ("Token", "code"),
        ("Nhóm", "group_name"),
        ("Vai trò", "role"),
        ("Trạng thái", "status"),
        ("Họ tên", "owner_name"),
        ("Email", "owner_email"),
        ("Ngày tạo", "created_at"),
        ("Hết hạn", "expires_at"),
        ("Ngày kích hoạt", "redeemed_at"),
        ("Thiết bị", "device_count"),
        ("Ghi chú", "export_note"),
    ]

    def add_sheet(title: str, items: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet(title)
        sheet.append([label for label, _ in columns])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="245B86")
        for index, row in enumerate(items, start=1):
            values = {**row, "index": index}
            sheet.append(
                [_excel_safe(values.get(key, "") or "") for _, key in columns]
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, sheet.max_row)}"
        widths = [8, 28, 24, 14, 16, 24, 30, 20, 20, 20, 12, 34]
        for column_index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    available = [
        row
        for row in all_rows
        if row.get("status_key") == "available" and row.get("exportable")
    ]
    add_sheet("Token chưa dùng", available)
    add_sheet("Tất cả token", all_rows)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
