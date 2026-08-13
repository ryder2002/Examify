"""Integration tests for activation, My Exams and attempts using SQLite."""

from __future__ import annotations

import os
import tempfile
import unittest
import importlib
import hashlib
from datetime import timedelta
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import Mock, patch


_database_file = tempfile.NamedTemporaryFile(prefix="smart-exam-test-", suffix=".db", delete=False)
_database_file.close()
os.environ["DATABASE_URL"] = f"sqlite+pysqlite:///{_database_file.name}"
os.environ["AUTH_REQUIRED"] = "true"
os.environ["ADMIN_EMAIL"] = "admin@test.local"
os.environ["ADMIN_PASSWORD"] = "test-admin-password"
os.environ["JWT_SECRET"] = "test-secret-that-is-long-enough-for-local-integration-tests"
os.environ["TOKEN_EXPORT_SECRET"] = "test-token-export-secret-that-is-long-enough"

from fastapi import HTTPException  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import event, func, select  # noqa: E402

import auth_service  # noqa: E402
import config  # noqa: E402
import database  # noqa: E402
import models  # noqa: E402

importlib.reload(config)
importlib.reload(database)
importlib.reload(models)
importlib.reload(auth_service)

from auth_service import bootstrap_admin, hash_password, sha256  # noqa: E402
from database import session_scope  # noqa: E402
from full_test_components import abandon_pending_components  # noqa: E402
from main import app  # noqa: E402
import desktop_sync_api  # noqa: E402
import maintenance_tasks  # noqa: E402
import platform_api  # noqa: E402
from models import (  # noqa: E402
    Asset,
    DesktopSync,
    Device,
    Exam,
    ExamSource,
    Job,
    QuestionRecord,
    SitePolicy,
    User,
    utcnow,
    uuid4,
)
from platform_api import persist_final_exam  # noqa: E402
from rate_limit import rate_limiter  # noqa: E402


class PlatformApiTests(unittest.TestCase):
    PASSWORD = "test-user-password"

    @classmethod
    def setUpClass(cls) -> None:
        database.create_schema()
        bootstrap_admin()

    @classmethod
    def tearDownClass(cls) -> None:
        if database.engine is not None:
            database.engine.dispose()
        Path(_database_file.name).unlink(missing_ok=True)

    def setUp(self) -> None:
        rate_limiter.reset_local_for_tests()

    def register_and_login(
        self,
        client: TestClient,
        *,
        email: str,
        device_key: str,
        display_name: str = "Test User",
    ) -> None:
        client.headers.update({"X-Examify-Device-Key": device_key})
        registered = client.post(
            "/api/v1/auth/register",
            json={
                "display_name": display_name,
                "email": email,
                "password": self.PASSWORD,
                "password_confirmation": self.PASSWORD,
            },
        )
        self.assertEqual(registered.status_code, 200, registered.text)
        # Registration deliberately does not grant an application session.
        self.assertEqual(client.get("/api/v1/auth/me").status_code, 401)
        logged_in = client.post(
            "/api/v1/auth/login",
            json={
                "email": email,
                "password": self.PASSWORD,
                "device_key": device_key,
            },
        )
        self.assertEqual(logged_in.status_code, 200, logged_in.text)

    def desktop_register_and_login(
        self,
        client: TestClient,
        *,
        activation: dict,
        email: str,
        device_key: str,
    ) -> dict:
        registered = client.post(
            "/api/v1/desktop/auth/register",
            json={
                "display_name": "Desktop User",
                "email": email,
                "password": self.PASSWORD,
                "password_confirmation": self.PASSWORD,
                "setup_token": activation["setup_token"],
            },
        )
        self.assertEqual(registered.status_code, 200, registered.text)
        logged_in = client.post(
            "/api/v1/desktop/auth/login",
            json={
                "email": email,
                "password": self.PASSWORD,
                "device_key": device_key,
            },
        )
        self.assertEqual(logged_in.status_code, 200, logged_in.text)
        return logged_in.json()

    def test_persist_final_exam_updates_same_desktop_exam_without_consuming_new_id(self) -> None:
        with session_scope() as session:
            owner_id = session.scalar(
                select(User.id).where(User.email == config.settings.admin_email)
            )
        initial = {
            "schema_version": 2,
            "exam_type": "reading",
            "questions": [{"number": 101, "correct": "A"}],
            "stimuli": [],
            "audios": [],
        }
        exam_id = persist_final_exam(
            initial,
            job_id=None,
            owner_user_id=owner_id,
            title="Desktop mutable exam",
            category="Old",
        )
        updated = {
            **initial,
            "questions": [
                {"number": 101, "correct": "B"},
                {"number": 102, "correct": "C"},
            ],
        }
        updated_id = persist_final_exam(
            updated,
            job_id=None,
            owner_user_id=owner_id,
            title="Desktop mutable exam v2",
            category="New",
            target_exam_id=exam_id,
        )

        self.assertEqual(updated_id, exam_id)
        with session_scope() as session:
            exam = session.get(Exam, exam_id)
            question_count = session.scalar(
                select(func.count(QuestionRecord.id)).where(
                    QuestionRecord.exam_id == exam_id
                )
            )
            self.assertEqual(exam.title, "Desktop mutable exam v2")
            self.assertEqual(exam.category, "New")
            self.assertEqual(exam.question_count, 2)
            self.assertEqual(question_count, 2)

    def test_abandon_reading_component_refunds_reserved_quota_once(self) -> None:
        owner_id = uuid4()
        component_id = uuid4()
        with session_scope() as session:
            session.add(
                User(
                    id=owner_id,
                    email=f"quota-refund-{owner_id}@test.local",
                    display_name="Quota refund teacher",
                    role="teacher",
                    status="active",
                    exam_limit=1,
                    exam_created_count=1,
                )
            )
            session.add(
                Exam(
                    id=component_id,
                    owner_user_id=owner_id,
                    title="Reading staging",
                    exam_type="reading",
                    status="component_pending",
                    library_scope="personal",
                    question_count=100,
                    payload={},
                )
            )

        with session_scope() as session:
            abandoned = abandon_pending_components(
                session,
                owner_user_id=owner_id,
                exam_ids={component_id},
            )
            self.assertEqual(abandoned, [component_id])
        with session_scope() as session:
            self.assertEqual(session.get(User, owner_id).exam_created_count, 0)
            self.assertEqual(
                session.get(Exam, component_id).status,
                "component_abandoned",
            )
            self.assertEqual(
                abandon_pending_components(
                    session,
                    owner_user_id=owner_id,
                    exam_ids={component_id},
                ),
                [],
            )
            self.assertEqual(session.get(User, owner_id).exam_created_count, 0)

    def test_full_test_components_stay_out_of_shared_bank_until_combined(self) -> None:
        with session_scope() as session:
            owner_id = session.scalar(
                select(User.id).where(User.email == config.settings.admin_email)
            )
        listening_id = persist_final_exam(
            {
                "schema_version": 2,
                "job_id": "staged-listening-job",
                "exam_type": "listening",
                "questions": [
                    {"number": number, "correct": "A"}
                    for number in range(1, 101)
                ],
                "stimuli": [],
                "audios": [],
            },
            job_id=None,
            owner_user_id=owner_id,
            title="Listening Component",
        )
        reading_id = persist_final_exam(
            {
                "schema_version": 2,
                "job_id": "staged-reading-job",
                "exam_type": "reading",
                "questions": [
                    {"number": number, "correct": "B"}
                    for number in range(101, 201)
                ],
                "stimuli": [],
                "audios": [],
            },
            job_id=None,
            owner_user_id=owner_id,
            title="Full Test staging regression",
            is_full_test_component=True,
        )

        with session_scope() as session:
            for component_id in (listening_id, reading_id):
                component = session.get(Exam, component_id)
                self.assertEqual(component.library_scope, "personal")
                self.assertIsNone(component.shared_title_key)
                self.assertEqual(component.status, "component_pending")

        with TestClient(app) as client:
            logged_in = client.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-full-test-component-key",
                },
            )
            self.assertEqual(logged_in.status_code, 200, logged_in.text)
            personal_before_combine = client.get("/api/v1/exams")
            self.assertEqual(
                personal_before_combine.status_code,
                200,
                personal_before_combine.text,
            )
            self.assertTrue(
                {listening_id, reading_id}.isdisjoint(
                    {item["id"] for item in personal_before_combine.json()["items"]}
                )
            )
            combined = client.post(
                "/api/v1/exams/combine",
                json={
                    "listening_exam_id": listening_id,
                    "reading_exam_id": reading_id,
                    "title": "Full Test staging regression",
                    "category": "ETS regression",
                },
            )
            self.assertEqual(combined.status_code, 200, combined.text)
            self.assertEqual(combined.json()["exam_type"], "combined")
            self.assertEqual(len(combined.json()["questions"]), 200)
            combined_id = combined.json()["exam_id"]
            retried = client.post(
                "/api/v1/exams/combine",
                json={
                    "listening_exam_id": listening_id,
                    "reading_exam_id": reading_id,
                    "title": "Full Test staging regression",
                    "category": "ETS regression",
                },
            )
            self.assertEqual(retried.status_code, 200, retried.text)
            self.assertEqual(retried.json()["exam_id"], combined_id)

        with session_scope() as session:
            stored = session.get(Exam, combined_id)
            self.assertEqual(stored.library_scope, "teacher_shared")
            self.assertEqual(stored.answer_key_count, 200)
            self.assertIsNotNone(session.get(Exam, listening_id).deleted_at)
            self.assertIsNotNone(session.get(Exam, reading_id).deleted_at)
            self.assertEqual(
                session.get(Exam, listening_id).status,
                "combined_component",
            )
            self.assertEqual(
                session.get(Exam, reading_id).status,
                "combined_component",
            )

    def test_finalize_allows_multiple_pending_listening_components(self) -> None:
        with session_scope() as session:
            owner_id = session.scalar(
                select(User.id).where(User.email == config.settings.admin_email)
            )
            job_id = uuid4()
            session.add(
                Job(
                    id=job_id,
                    owner_user_id=owner_id,
                    exam_type="listening",
                    filename="second-listening.pdf",
                    file_hash=hashlib.sha256(job_id.encode()).hexdigest(),
                    pipeline_version="test",
                    status="review",
                    payload={},
                )
            )

        questions = [
            {
                "number": number,
                "part": (
                    "Part 1"
                    if number <= 6
                    else "Part 2"
                    if number <= 31
                    else "Part 3"
                ),
                "text": "" if number <= 31 else f"Question {number}",
                "options": (
                    {}
                    if number <= 31
                    else {"A": "One", "B": "Two", "C": "Three", "D": "Four"}
                ),
                "option_letters": (
                    ["A", "B", "C"] if 7 <= number <= 31 else ["A", "B", "C", "D"]
                ),
                "correct": "A",
                "issues": [],
            }
            for number in range(1, 101)
        ]
        old_component_id = persist_final_exam(
            {
                "schema_version": 2,
                "job_id": "older-unfinished-listening",
                "exam_type": "listening",
                "questions": questions,
                "stimuli": [],
                "audios": [],
            },
            job_id=None,
            owner_user_id=owner_id,
            title="Listening Component",
            is_full_test_component=True,
        )
        state = {
            "job_id": job_id,
            "exam_type": "listening",
            "filename": "second-listening.pdf",
            "status": "review",
            "stage": "Review",
            "questions": questions,
            "stimuli": [],
            "solutions": [],
            "audios": [],
            "audio": None,
        }
        fake_store = SimpleNamespace(
            read=lambda _job_id: state,
            write=lambda _job_id, updated: state.update(updated),
            owner_id=lambda _job_id: owner_id,
            cleanup=lambda: None,
        )

        with patch("main.store", fake_store), TestClient(app) as client:
            logged_in = client.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": f"pending-component-{job_id}",
                },
            )
            self.assertEqual(logged_in.status_code, 200, logged_in.text)
            finalized = client.post(
                f"/api/extractions/{job_id}/finalize",
                json={
                    "answer_key": {},
                    "title": "Listening Component",
                    "category": "",
                    "is_full_test_component": True,
                },
            )
            retried = client.post(
                f"/api/extractions/{job_id}/finalize",
                json={
                    "answer_key": {},
                    "title": "Listening Component",
                    "category": "",
                    "is_full_test_component": True,
                },
            )
            self.assertEqual(finalized.status_code, 200, finalized.text)
            self.assertEqual(retried.status_code, 200, retried.text)
            component_id = finalized.json()["exam_id"]
            self.assertEqual(retried.json()["exam_id"], component_id)
            self.assertNotEqual(component_id, old_component_id)
            with session_scope() as session:
                components = session.scalars(
                    select(Exam).where(
                        Exam.owner_user_id == owner_id,
                        Exam.title == "Listening Component",
                        Exam.status == "component_pending",
                        Exam.deleted_at.is_(None),
                    )
                ).all()
                self.assertGreaterEqual(len(components), 2)

            cancelled = client.delete(
                f"/api/v1/full-test-components/{component_id}"
            )
            cancelled_again = client.delete(
                f"/api/v1/full-test-components/{component_id}"
            )
            self.assertEqual(cancelled.status_code, 200, cancelled.text)
            self.assertEqual(cancelled_again.status_code, 200, cancelled_again.text)
            with session_scope() as session:
                self.assertEqual(
                    session.get(Exam, component_id).status,
                    "component_abandoned",
                )
                self.assertEqual(
                    session.get(Exam, old_component_id).status,
                    "component_pending",
                )

            logged_out = client.post("/api/v1/auth/logout")
            self.assertEqual(logged_out.status_code, 200, logged_out.text)

        with session_scope() as session:
            for component_id in (old_component_id, component_id):
                component = session.get(Exam, component_id)
                self.assertEqual(component.status, "component_abandoned")
                self.assertIsNotNone(component.deleted_at)

    def test_purge_stale_full_test_component_removes_storage_and_database(self) -> None:
        owner_id = uuid4()
        job_id = uuid4()
        component_id = uuid4()
        source_key = f"examify-sources/{component_id}/listening.pdf"
        old = utcnow() - timedelta(days=2)
        with session_scope() as session:
            session.add(
                User(
                    id=owner_id,
                    email=f"stale-component-{owner_id}@test.local",
                    display_name="Stale component teacher",
                    role="teacher",
                    status="active",
                    exam_limit=1,
                    exam_created_count=0,
                )
            )
            session.add(
                Job(
                    id=job_id,
                    owner_user_id=owner_id,
                    exam_type="listening",
                    filename="stale-listening.pdf",
                    file_hash=hashlib.sha256(job_id.encode()).hexdigest(),
                    pipeline_version="test",
                    status="ready",
                    payload={},
                    updated_at=old,
                )
            )
            session.add(
                Exam(
                    id=component_id,
                    owner_user_id=owner_id,
                    job_id=job_id,
                    title="Listening Component",
                    exam_type="listening",
                    status="component_pending",
                    library_scope="personal",
                    question_count=100,
                    payload={},
                    updated_at=old,
                )
            )
            session.flush()
            session.add(
                Asset(
                    exam_id=component_id,
                    kind="audio",
                    bucket="examify-audio",
                    object_key=f"jobs/{job_id}/audio/full.mp3",
                    filename="full.mp3",
                    content_type="audio/mpeg",
                    size=100,
                )
            )
            session.add(
                ExamSource(
                    exam_id=component_id,
                    component="listening",
                    bucket="examify-sources",
                    object_key=source_key,
                    filename="listening.pdf",
                )
            )

        removed_objects: list[tuple[str, str]] = []
        removed_prefixes: list[tuple[str, str]] = []
        fake_storage = SimpleNamespace(
            remove_object=lambda bucket, key: removed_objects.append((bucket, key)),
            remove_prefix=lambda bucket, prefix: removed_prefixes.append(
                (bucket, prefix)
            ),
        )
        with patch.object(maintenance_tasks, "storage", fake_storage):
            result = maintenance_tasks.purge_full_test_components(limit=100)

        self.assertGreaterEqual(result["stale_abandoned"], 1)
        self.assertIn(("examify-sources", source_key), removed_objects)
        self.assertIn(("examify-audio", f"jobs/{job_id}/"), removed_prefixes)
        with session_scope() as session:
            self.assertIsNone(session.get(Exam, component_id))
            self.assertIsNone(session.get(Job, job_id))

    def test_combine_reuses_sources_created_during_final_persist(self) -> None:
        """Combining source-backed components must not insert duplicate sources."""

        class FakeStorage:
            def __init__(self) -> None:
                self.copies: list[tuple[str, str, str]] = []

                class Client:
                    @staticmethod
                    def stat_object(bucket: str, object_key: str) -> SimpleNamespace:
                        return SimpleNamespace(size=1024)

                self.client = Client()

            def copy_object(
                self, bucket: str, source_key: str, destination_key: str
            ) -> None:
                self.copies.append((bucket, source_key, destination_key))

        with session_scope() as session:
            owner_id = session.scalar(
                select(User.id).where(User.email == config.settings.admin_email)
            )
        fake_storage = FakeStorage()
        listening_job_id = uuid4()
        reading_job_id = uuid4()

        with patch.object(platform_api, "storage", fake_storage):
            listening_id = persist_final_exam(
                {
                    "schema_version": 2,
                    "job_id": listening_job_id,
                    "exam_type": "listening",
                    "questions": [
                        {"number": number, "correct": "A"}
                        for number in range(1, 101)
                    ],
                    "stimuli": [],
                    "audios": [],
                },
                job_id=listening_job_id,
                owner_user_id=owner_id,
                title="Listening Component",
            )
            reading_id = persist_final_exam(
                {
                    "schema_version": 2,
                    "job_id": reading_job_id,
                    "exam_type": "reading",
                    "questions": [
                        {"number": number, "correct": "B"}
                        for number in range(101, 201)
                    ],
                    "stimuli": [],
                    "audios": [],
                },
                job_id=reading_job_id,
                owner_user_id=owner_id,
                title="Source-backed Reading Component",
                is_full_test_component=True,
            )
            copies_before_combine = len(fake_storage.copies)

            with TestClient(app) as client:
                logged_in = client.post(
                    "/api/v1/auth/login",
                    json={
                        "email": "admin@test.local",
                        "password": "test-admin-password",
                        "device_key": "admin-source-backed-combine-key",
                    },
                )
                self.assertEqual(logged_in.status_code, 200, logged_in.text)
                response = client.post(
                    "/api/v1/exams/combine",
                    json={
                        "listening_exam_id": listening_id,
                        "reading_exam_id": reading_id,
                        "title": "Source-backed Full Test",
                        "category": "ETS regression",
                    },
                )
            self.assertEqual(response.status_code, 200, response.text)
            combined_id = response.json()["exam_id"]

            with session_scope() as session:
                sources = list(
                    session.scalars(
                        select(ExamSource).where(ExamSource.exam_id == combined_id)
                    )
                )
                self.assertEqual(
                    {source.component for source in sources},
                    {"listening", "reading"},
                )
                self.assertEqual(len(sources), 2)
                self.assertEqual(
                    {
                        source.object_key
                        for source in sources
                    },
                    {
                        f"examify-sources/{combined_id}/listening.pdf",
                        f"examify-sources/{combined_id}/reading.pdf",
                    },
                )

            # Two component copies plus two copies made by persist_final_exam
            # for the combined payload.  The second combine phase must reuse
            # those rows rather than copying/inserting them again.
            self.assertEqual(len(fake_storage.copies), copies_before_combine + 2)

    def test_combine_retry_repairs_partially_committed_combined_exam(self) -> None:
        """A retry cleans a combined row whose asset move failed after commit."""

        with session_scope() as session:
            owner_id = session.scalar(
                select(User.id).where(User.email == config.settings.admin_email)
            )
        listening_job_id = uuid4()
        reading_job_id = uuid4()
        listening_id = persist_final_exam(
            {
                "schema_version": 2,
                "job_id": listening_job_id,
                "exam_type": "listening",
                "questions": [
                    {"number": number, "correct": "A"}
                    for number in range(1, 101)
                ],
                "stimuli": [],
                "audios": [],
            },
            job_id=listening_job_id,
            owner_user_id=owner_id,
            title="Listening Component",
        )
        reading_id = persist_final_exam(
            {
                "schema_version": 2,
                "job_id": reading_job_id,
                "exam_type": "reading",
                "questions": [
                    {"number": number, "correct": "B"}
                    for number in range(101, 201)
                ],
                "stimuli": [],
                "audios": [],
            },
            job_id=reading_job_id,
            owner_user_id=owner_id,
            title="Partially Committed Reading Component",
            is_full_test_component=True,
        )
        combined_id = persist_final_exam(
            {
                "schema_version": 2,
                "job_id": f"{listening_job_id}+{reading_job_id}",
                "component_job_ids": {
                    "listening": listening_job_id,
                    "reading": reading_job_id,
                },
                "exam_type": "combined",
                "questions": [
                    {"number": number, "correct": "A" if number < 101 else "B"}
                    for number in range(1, 201)
                ],
                "stimuli": [],
                "audios": [],
            },
            job_id=None,
            owner_user_id=owner_id,
            title="Recoverable Full Test",
            category="Recovery regression",
            quota_replacement_exam_ids=(listening_id, reading_id),
            defer_version_snapshot=True,
        )

        with session_scope() as session:
            self.assertIsNone(session.get(Exam, combined_id).current_version_id)
            self.assertEqual(session.get(Exam, listening_id).status, "component_pending")

        with TestClient(app) as client:
            logged_in = client.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-partial-combine-retry-key",
                },
            )
            self.assertEqual(logged_in.status_code, 200, logged_in.text)
            response = client.post(
                "/api/v1/exams/combine",
                json={
                    "listening_exam_id": listening_id,
                    "reading_exam_id": reading_id,
                    "title": "Recoverable Full Test",
                    "category": "Recovery regression",
                },
            )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["exam_id"], combined_id)

        with session_scope() as session:
            self.assertIsNotNone(session.get(Exam, combined_id).current_version_id)
            for component_id in (listening_id, reading_id):
                component = session.get(Exam, component_id)
                self.assertEqual(component.status, "combined_component")
                self.assertIsNotNone(component.deleted_at)

    def test_combined_exam_can_replace_legacy_shared_component_title(self) -> None:
        with session_scope() as session:
            owner_id = session.scalar(
                select(User.id).where(User.email == config.settings.admin_email)
            )
        legacy_id = persist_final_exam(
            {
                "schema_version": 2,
                "exam_type": "reading",
                "questions": [{"number": 101, "correct": "A"}],
                "stimuli": [],
                "audios": [],
            },
            job_id=None,
            owner_user_id=owner_id,
            title="Legacy shared component replacement",
        )
        combined_id = persist_final_exam(
            {
                "schema_version": 2,
                "exam_type": "combined",
                "questions": [
                    {"number": 1, "correct": "B"},
                    {"number": 101, "correct": "A"},
                ],
                "stimuli": [],
                "audios": [],
            },
            job_id=None,
            owner_user_id=owner_id,
            title="Legacy shared component replacement",
            quota_replacement_exam_ids=(legacy_id,),
        )

        with session_scope() as session:
            self.assertIsNone(session.get(Exam, legacy_id).shared_title_key)
            self.assertIsNotNone(session.get(Exam, combined_id).shared_title_key)

    def test_desktop_sync_reconciles_changed_manifest_into_same_remote_exam(self) -> None:
        class FakeStorage:
            def __init__(self) -> None:
                self.objects: dict[tuple[str, str], bytes] = {}

            def put_file(self, bucket, object_key, source, content_type=None):
                self.objects[(bucket, object_key)] = Path(source).read_bytes()

        fake_storage = FakeStorage()
        asset_bytes = b"desktop-sync-asset"
        with session_scope() as session:
            user = User(
                email="desktop-reconcile@test.local",
                display_name="Desktop Reconcile",
                password_hash=hash_password(self.PASSWORD),
                registered_at=utcnow(),
                role="user",
                status="active",
            )
            session.add(user)
            session.flush()
            session.add(
                Device(
                    user_id=user.id,
                    device_key_hash=sha256("desktop-sync-device-key-0001"),
                    identity_kind="legacy_browser",
                    name="Desktop sync test",
                    platform="test",
                )
            )
        with patch.object(desktop_sync_api, "storage", fake_storage), TestClient(app) as client:
            client.headers.update(
                {"X-Examify-Device-Key": "desktop-sync-device-key-0001"}
            )
            login = client.post(
                "/api/v1/auth/login",
                json={
                    "email": "desktop-reconcile@test.local",
                    "password": self.PASSWORD,
                    "device_key": "desktop-sync-device-key-0001",
                },
            )
            self.assertEqual(login.status_code, 200, login.text)
            client.headers.update({"X-Examify-Desktop-Version": "0.1.3"})
            epoch = client.get("/api/v1/system/data-epoch")
            self.assertEqual(epoch.status_code, 200, epoch.text)
            manifest = {
                "data_epoch": epoch.json()["data_epoch"],
                "client_exam_id": "desktop-reconcile-exam-0001",
                "title": "Desktop reconcile v1",
                "category": "ETS",
                "payload": {
                    "schema_version": 2,
                    "exam_type": "reading",
                    "questions": [{"number": 101, "correct": "A"}],
                    "stimuli": [
                        {
                            "id": "stimulus-101",
                            "assets": [
                                {
                                    "id": "question-101.webp",
                                    "filename": "question-101.webp",
                                    "url": "/local",
                                }
                            ],
                        }
                    ],
                    "audios": [],
                },
                "assets": [
                    {
                        "asset_id": "question-101.webp",
                        "kind": "stimulus",
                        "filename": "question-101.webp",
                        "content_type": "image/webp",
                        "size": len(asset_bytes),
                        "sha256": hashlib.sha256(asset_bytes).hexdigest(),
                    }
                ],
            }
            stale_manifest = {**manifest, "data_epoch": "00000000-0000-4000-8000-000000000000"}
            stale = client.post("/api/v1/desktop/sync/exams", json=stale_manifest)
            self.assertEqual(stale.status_code, 409, stale.text)
            self.assertEqual(stale.json()["detail"]["code"], "stale_data_epoch")
            created = client.post("/api/v1/desktop/sync/exams", json=manifest)
            self.assertEqual(created.status_code, 200, created.text)
            premature = client.post(
                f"/api/v1/desktop/sync/exams/{created.json()['sync_id']}/complete"
            )
            self.assertEqual(premature.status_code, 409, premature.text)
            wrong_asset = client.put(
                f"/api/v1/desktop/sync/exams/{created.json()['sync_id']}"
                "/assets/question-101.webp",
                files={"file": ("question-101.webp", b"wrong", "image/webp")},
            )
            self.assertEqual(wrong_asset.status_code, 422, wrong_asset.text)
            uploaded_asset = client.put(
                f"/api/v1/desktop/sync/exams/{created.json()['sync_id']}"
                "/assets/question-101.webp",
                files={"file": ("question-101.webp", asset_bytes, "image/webp")},
            )
            self.assertEqual(uploaded_asset.status_code, 200, uploaded_asset.text)
            completed = client.post(
                f"/api/v1/desktop/sync/exams/{created.json()['sync_id']}/complete"
            )
            self.assertEqual(completed.status_code, 200, completed.text)
            self.assertEqual(list(fake_storage.objects.values()), [asset_bytes])
            exam_id = completed.json()["exam_id"]

            changed = {
                **manifest,
                "base_revision": completed.json()["revision"],
                "title": "Desktop reconcile v2",
                "payload": {
                    **manifest["payload"],
                    "questions": [
                        {"number": 101, "correct": "B"},
                        {"number": 102, "correct": "C"},
                    ],
                },
            }
            restarted = client.post("/api/v1/desktop/sync/exams", json=changed)
            self.assertEqual(restarted.status_code, 200, restarted.text)
            self.assertEqual(restarted.json()["status"], "uploading")
            reconciled = client.post(
                f"/api/v1/desktop/sync/exams/{restarted.json()['sync_id']}/complete"
            )
            self.assertEqual(reconciled.status_code, 200, reconciled.text)
            self.assertEqual(reconciled.json()["exam_id"], exam_id)
            self.assertEqual(reconciled.json()["revision"], 2)

            receipt = client.get("/api/v1/desktop/sync/reconcile")
            self.assertEqual(receipt.status_code, 200, receipt.text)
            self.assertIn(
                {
                    "client_exam_id": manifest["client_exam_id"],
                    "exam_id": exam_id,
                    "revision": 2,
                    "deleted": False,
                },
                receipt.json()["items"],
            )

            stale_edit = {
                **changed,
                "title": "Desktop stale overwrite",
                "payload": {
                    **changed["payload"],
                    "questions": [{"number": 101, "correct": "D"}],
                },
            }
            conflict = client.post("/api/v1/desktop/sync/exams", json=stale_edit)
            self.assertEqual(conflict.status_code, 409, conflict.text)
            self.assertEqual(
                conflict.json()["detail"]["code"], "exam_revision_conflict"
            )

            duplicate = client.post(
                f"/api/v1/desktop/sync/exams/{restarted.json()['sync_id']}/complete"
            )
            self.assertEqual(duplicate.status_code, 200, duplicate.text)
            self.assertEqual(duplicate.json()["exam_id"], exam_id)

            # Model a worker crash after persist_final_exam committed but before
            # the sync receipt was updated. A stale completion lease must reuse
            # the exam protected by (owner, client_exam_id), not create another.
            with session_scope() as session:
                sync_row = session.get(DesktopSync, restarted.json()["sync_id"])
                sync_row.status = "completing"
                sync_row.exam_id = None
                sync_row.updated_at = utcnow().replace(year=2020)
            recovered = client.post(
                f"/api/v1/desktop/sync/exams/{restarted.json()['sync_id']}/complete"
            )
            self.assertEqual(recovered.status_code, 200, recovered.text)
            self.assertEqual(recovered.json()["exam_id"], exam_id)

        with session_scope() as session:
            exam = session.get(Exam, exam_id)
            exam_count = session.scalar(
                select(func.count(Exam.id)).where(
                    Exam.owner_user_id == exam.owner_user_id,
                    Exam.client_exam_id == manifest["client_exam_id"],
                )
            )
            self.assertEqual(exam.title, "Desktop reconcile v2")
            self.assertEqual(exam.question_count, 2)
            self.assertEqual(exam_count, 1)

    def test_bulk_student_tokens_groups_and_repeatable_excel_export(self) -> None:
        with TestClient(app) as admin:
            login = admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-bulk-token-device-key-0001",
                },
            )
            self.assertEqual(login.status_code, 200, login.text)
            center_a = admin.post(
                "/api/v1/admin/token-groups", json={"name": "Trung Tâm A"}
            )
            self.assertEqual(center_a.status_code, 200, center_a.text)
            center_a_id = center_a.json()["id"]
            created = admin.post(
                "/api/v1/admin/tokens",
                json={
                    "count": 200,
                    "label": "Nhãn phải bị bỏ qua",
                    "assigned_role": "student",
                    "group_id": center_a_id,
                },
            )
            self.assertEqual(created.status_code, 200, created.text)
            codes = created.json()["codes"]
            self.assertEqual(len(codes), 200)
            self.assertEqual(len(set(codes)), 200)
            self.assertRegex(
                codes[0],
                r"^EXAMIFY-[A-Z2-9]{4}(?:-[A-Z2-9]{4}){3}$",
            )

            page_one = admin.get(
                "/api/v1/admin/tokens",
                params={"group_id": center_a_id, "page": 1, "page_size": 100},
            )
            self.assertEqual(page_one.status_code, 200, page_one.text)
            self.assertEqual(page_one.json()["total"], 200)
            self.assertEqual(page_one.json()["pages"], 2)
            self.assertTrue(all(item["group_id"] == center_a_id for item in page_one.json()["items"]))
            self.assertTrue(all(item["label"] == "" for item in page_one.json()["items"]))
            self.assertTrue(all("code" not in item for item in page_one.json()["items"]))

            first_export = admin.get(
                f"/api/v1/admin/token-groups/{center_a_id}/export.xlsx"
            )
            self.assertEqual(first_export.status_code, 200, first_export.text)
            workbook = load_workbook(BytesIO(first_export.content), read_only=True)
            self.assertEqual(workbook.sheetnames, ["Token chưa dùng", "Tất cả token"])
            unused_codes = {
                row[1]
                for row in workbook["Token chưa dùng"].iter_rows(min_row=2, values_only=True)
            }
            self.assertEqual(unused_codes, set(codes))
            self.assertEqual(workbook["Tất cả token"].max_row, 201)

            with session_scope() as session:
                stored = session.scalars(
                    select(models.ActivationToken).where(
                        models.ActivationToken.group_id == center_a_id
                    )
                ).all()
                self.assertEqual(len(stored), 200)
                self.assertTrue(all(item.encrypted_code for item in stored))
                self.assertTrue(
                    all(code not in item.encrypted_code for code, item in zip(codes, stored))
                )

        with TestClient(app) as student:
            student.headers.update(
                {"X-Examify-Device-Key": "bulk-student-device-key-0000001"}
            )
            redeemed = student.post(
                "/api/v1/activations/redeem",
                json={
                    "code": codes[0],
                    "device_key": "bulk-student-device-key-0000001",
                },
            )
            self.assertEqual(redeemed.status_code, 200, redeemed.text)
            registered = student.post(
                "/api/v1/auth/register",
                json={
                    "display_name": "Học Viên Trung Tâm A",
                    "email": "bulk-student@test.local",
                    "password": self.PASSWORD,
                    "password_confirmation": self.PASSWORD,
                },
            )
            self.assertEqual(registered.status_code, 200, registered.text)

        with TestClient(app) as admin:
            admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-bulk-token-device-key-0001",
                },
            )
            named = admin.get(
                "/api/v1/admin/tokens",
                params={"group_id": center_a_id, "search": "Học Viên Trung Tâm A"},
            )
            self.assertEqual(named.status_code, 200, named.text)
            self.assertEqual(named.json()["total"], 1)
            self.assertEqual(named.json()["items"][0]["label"], "Học Viên Trung Tâm A")
            self.assertEqual(named.json()["items"][0]["owner_email"], "bulk-student@test.local")

            repeated_export = admin.get(
                f"/api/v1/admin/token-groups/{center_a_id}/export.xlsx"
            )
            repeated = load_workbook(BytesIO(repeated_export.content), read_only=True)
            self.assertEqual(repeated["Token chưa dùng"].max_row, 200)
            self.assertEqual(repeated["Tất cả token"].max_row, 201)
            all_export_rows = list(
                repeated["Tất cả token"].iter_rows(min_row=2, values_only=True)
            )
            self.assertTrue(
                any(
                    row[1] == codes[0]
                    and row[5] == "Học Viên Trung Tâm A"
                    and row[6] == "bulk-student@test.local"
                    and row[4] == "Đã kích hoạt"
                    for row in all_export_rows
                )
            )
            with session_scope() as session:
                export_audits = session.scalars(
                    select(models.AuditLog).where(
                        models.AuditLog.action == "activation_group.exported",
                        models.AuditLog.target_id == center_a_id,
                    )
                ).all()
                self.assertEqual(len(export_audits), 2)

            center_b = admin.post(
                "/api/v1/admin/token-groups", json={"name": "Trung Tâm B"}
            )
            self.assertEqual(center_b.status_code, 200, center_b.text)
            center_b_id = center_b.json()["id"]
            moving_ids = [item["id"] for item in page_one.json()["items"][:2]]
            moved = admin.patch(
                "/api/v1/admin/tokens/group-membership",
                json={"token_ids": moving_ids, "group_id": center_b_id},
            )
            self.assertEqual(moved.status_code, 200, moved.text)
            renamed = admin.patch(
                f"/api/v1/admin/token-groups/{center_b_id}",
                json={"name": "Trung Tâm B Mới"},
            )
            self.assertEqual(renamed.status_code, 200, renamed.text)
            self.assertEqual(renamed.json()["total"], 2)
            deleted = admin.delete(f"/api/v1/admin/token-groups/{center_b_id}")
            self.assertEqual(deleted.status_code, 200, deleted.text)
            self.assertEqual(deleted.json()["tokens_moved_to_ungrouped"], 2)
            ungrouped = admin.get(
                "/api/v1/admin/tokens",
                params={"group_id": "ungrouped", "page_size": 100},
            )
            ungrouped_ids = {item["id"] for item in ungrouped.json()["items"]}
            self.assertTrue(set(moving_ids).issubset(ungrouped_ids))

            bulk_deleted = admin.request(
                "DELETE",
                "/api/v1/admin/tokens",
                json={"token_ids": moving_ids},
            )
            self.assertEqual(bulk_deleted.status_code, 200, bulk_deleted.text)
            self.assertEqual(bulk_deleted.json()["deleted"], 2)
            remaining_ids = {
                item["id"]
                for item in admin.get(
                    "/api/v1/admin/tokens", params={"page_size": 100}
                ).json()["items"]
            }
            self.assertTrue(set(moving_ids).isdisjoint(remaining_ids))
            with session_scope() as session:
                bulk_audit = session.scalar(
                    select(models.AuditLog).where(
                        models.AuditLog.action
                        == "activation.bulk_permanently_deleted"
                    )
                )
                self.assertIsNotNone(bulk_audit)
                self.assertEqual(bulk_audit.detail["count"], 2)

    def test_non_admin_cannot_register_or_login_without_activation(self) -> None:
        with TestClient(app) as anonymous:
            registered = anonymous.post(
                "/api/v1/auth/register",
                json={
                    "display_name": "Not Activated",
                    "email": "not-activated@test.local",
                    "password": self.PASSWORD,
                    "password_confirmation": self.PASSWORD,
                },
            )
            self.assertEqual(registered.status_code, 401, registered.text)
            self.assertIn("kích hoạt", registered.json()["detail"].lower())

        # A pre-existing registered non-admin record cannot log in until
        # activation has created an active Device for it.
        with session_scope() as session:
            session.add(
                models.User(
                    display_name="No Device User",
                    email="no-device@test.local",
                    password_hash=auth_service.hash_password(self.PASSWORD),
                    registered_at=models.utcnow(),
                    role="user",
                )
            )
        with TestClient(app) as anonymous:
            logged_in = anonymous.post(
                "/api/v1/auth/login",
                json={
                    "email": "no-device@test.local",
                    "password": self.PASSWORD,
                    "device_key": "unactivated-device-key-000001",
                },
            )
            self.assertEqual(logged_in.status_code, 403, logged_in.text)
            self.assertIn("kích hoạt", logged_in.json()["detail"].lower())

    def test_one_time_activation_and_durable_attempt(self) -> None:
        with TestClient(app) as admin:
            login = admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-device-key-00000001",
                },
            )
            self.assertEqual(login.status_code, 200, login.text)
            created = admin.post(
                "/api/v1/admin/tokens",
                json={"count": 1, "label": "Test device", "expires_in_days": 30},
            )
            self.assertEqual(created.status_code, 200, created.text)
            code = created.json()["codes"][0]
            created_row = next(
                item
                for item in admin.get("/api/v1/admin/tokens").json()["items"]
                if item["label"] == "Test device"
            )
            self.assertIsNone(created_row["expires_at"])

        with TestClient(app) as user:
            activated = user.post(
                "/api/v1/activations/redeem",
                json={
                    "code": code,
                    "device_key": "user-device-key-0000000001",
                    "device_name": "Test Browser",
                },
            )
            self.assertEqual(activated.status_code, 200, activated.text)
            user_id = activated.json()["user_id"]
            self.register_and_login(
                client=user,
                email="integration-user@test.local",
                device_key="user-device-key-0000000001",
            )
            personal_tag = user.post(
                "/api/v1/tags",
                json={"name": "  ETS Personal 2026  "},
            )
            self.assertEqual(personal_tag.status_code, 200, personal_tag.text)
            self.assertEqual(personal_tag.json()["name"], "ETS Personal 2026")
            exam_payload = {
                "schema_version": 2,
                "job_id": "test-job",
                "exam_type": "reading",
                "requested_count": 1,
                "returned_count": 1,
                "total": 1,
                "questions": [
                    {
                        "number": 101,
                        "part": "Part 5",
                        "text": "Test question",
                        "options": {"A": "One", "B": "Two", "C": "Three", "D": "Four"},
                        "option_letters": ["A", "B", "C", "D"],
                        "correct": "A",
                        "group_id": None,
                        "stimulus_id": None,
                        "confidence": 100,
                        "issues": [],
                    }
                ],
                "stimuli": [],
                "audio": None,
                "audios": [],
                "solutions": [
                    {
                        "question_numbers": [101],
                        "transcript": None,
                        "explanation": "A is the correct answer.",
                        "translation": "A là đáp án đúng.",
                    }
                ],
            }
            exam_id = persist_final_exam(
                exam_payload,
                job_id=None,
                owner_user_id=user_id,
                title="Integration Exam",
            )
            listing = user.get("/api/v1/exams")
            self.assertEqual(listing.status_code, 200, listing.text)
            self.assertEqual(listing.json()["total"], 1)

            attempt = user.post(f"/api/v1/exams/{exam_id}/attempts", json={})
            self.assertEqual(attempt.status_code, 200, attempt.text)
            attempt_id = attempt.json()["attempt_id"]
            answer_statements: list[str] = []

            def capture_answer_sql(
                _connection, _cursor, statement, _parameters, _context, _many
            ) -> None:
                if "attempt_answers" in statement.lower():
                    answer_statements.append(statement.lower())

            event.listen(database.engine, "before_cursor_execute", capture_answer_sql)
            try:
                saved = user.patch(
                    f"/api/v1/attempts/{attempt_id}/answers",
                    json={
                        "answers": {"101": "A"},
                        "time_left_seconds": 11,
                        "client_revision": 1,
                    },
                )
            finally:
                event.remove(
                    database.engine, "before_cursor_execute", capture_answer_sql
                )
            self.assertEqual(saved.status_code, 200, saved.text)
            self.assertEqual(saved.json()["accepted_revision"], 1)
            self.assertEqual(
                sum(statement.lstrip().startswith("select") for statement in answer_statements),
                0,
                answer_statements,
            )
            stale = user.patch(
                f"/api/v1/attempts/{attempt_id}/answers",
                json={
                    "answers": {"101": "D"},
                    "time_left_seconds": 10,
                    "client_revision": 1,
                },
            )
            self.assertEqual(stale.status_code, 200, stale.text)
            self.assertEqual(stale.json()["accepted_revision"], 1)
            submit_statements: list[str] = []

            def capture_submit_sql(
                _connection, _cursor, statement, _parameters, _context, _many
            ) -> None:
                submit_statements.append(statement.lower())

            event.listen(database.engine, "before_cursor_execute", capture_submit_sql)
            try:
                submitted = user.post(
                    f"/api/v1/attempts/{attempt_id}/submit",
                    json={
                        "answers": {"101": "A"},
                        "time_left_seconds": 10,
                        "client_revision": 2,
                    },
                )
            finally:
                event.remove(
                    database.engine, "before_cursor_execute", capture_submit_sql
                )
            self.assertEqual(submitted.status_code, 200, submitted.text)
            self.assertLessEqual(len(submit_statements), 7, submit_statements)
            self.assertEqual(submitted.json()["answers"]["101"], "A")
            self.assertEqual(submitted.json()["accepted_revision"], 2)
            self.assertTrue(submitted.json()["has_solutions"])
            with database.SessionLocal() as verification_session:
                persisted_answer = verification_session.scalar(
                    select(models.AttemptAnswer).where(
                        models.AttemptAnswer.attempt_id == attempt_id,
                        models.AttemptAnswer.question_number == 101,
                    )
                )
                self.assertIs(persisted_answer.is_correct, True)
            duplicate_submit = user.post(
                f"/api/v1/attempts/{attempt_id}/submit",
                json={
                    "answers": {"101": "D"},
                    "time_left_seconds": 0,
                    "client_revision": 3,
                },
            )
            self.assertEqual(duplicate_submit.status_code, 200, duplicate_submit.text)
            self.assertEqual(duplicate_submit.json()["answers"]["101"], "A")
            self.assertEqual(duplicate_submit.json()["accepted_revision"], 2)
            attempt_detail = user.get(f"/api/v1/attempts/{attempt_id}")
            self.assertEqual(attempt_detail.status_code, 200, attempt_detail.text)
            self.assertTrue(attempt_detail.json()["has_solutions"])
            history = user.get(
                "/api/v1/attempts/history", params={"page": 1, "page_size": 1}
            )
            self.assertEqual(history.status_code, 200, history.text)
            self.assertEqual(history.json()["total"], 1)
            self.assertEqual(history.json()["page_size"], 1)
            self.assertEqual(len(history.json()["items"]), 1)
            self.assertEqual(history.json()["items"][0]["correct_count"], 1)
            self.assertTrue(history.json()["items"][0]["has_solutions"])

        with TestClient(app) as second_device:
            reused = second_device.post(
                "/api/v1/activations/redeem",
                json={
                    "code": code,
                    "device_key": "second-device-key-00000001",
                },
            )
            self.assertEqual(reused.status_code, 409, reused.text)

        # Admin can replace the device while retaining the same user/exams.
        with TestClient(app) as admin:
            admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-device-key-00000001",
                },
            )
            token_id = admin.get("/api/v1/admin/tokens").json()["items"][0]["id"]
            replacement = admin.post(
                f"/api/v1/admin/tokens/{token_id}/reissue",
                json={"revoke_existing_devices": True, "expires_in_days": 30},
            )
            self.assertEqual(replacement.status_code, 200, replacement.text)
            replacement_code = replacement.json()["code"]
            self.assertTrue(replacement.json()["data_preserved"])
            self.assertTrue(
                all(
                    item["expires_at"] is None
                    for item in admin.get("/api/v1/admin/tokens").json()["items"]
                )
            )

        with TestClient(app) as replacement_device:
            replacement_device.headers.update(
                {"X-Examify-Device-Key": "replacement-device-key-000001"}
            )
            activated = replacement_device.post(
                "/api/v1/activations/redeem",
                json={
                    "code": replacement_code,
                    "device_key": "replacement-device-key-000001",
                    "device_name": "Replacement Browser",
                },
            )
            self.assertEqual(activated.status_code, 200, activated.text)
            self.assertEqual(activated.json()["user_id"], user_id)
            replacement_device_id = activated.json()["device_id"]
            logged_in = replacement_device.post(
                "/api/v1/auth/login",
                json={
                    "email": "integration-user@test.local",
                    "password": self.PASSWORD,
                    "device_key": "replacement-device-key-000001",
                },
            )
            self.assertEqual(logged_in.status_code, 200, logged_in.text)
            listing = replacement_device.get("/api/v1/exams")
            self.assertEqual(listing.status_code, 200, listing.text)
            self.assertEqual(listing.json()["total"], 1)

        # The same operation is available from the Admin user-management view.
        with TestClient(app) as admin:
            admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-device-key-00000001",
                },
            )
            moved = admin.post(
                f"/api/v1/admin/users/{user_id}/reissue-token",
                json={"expires_in_days": 30},
            )
            self.assertEqual(moved.status_code, 200, moved.text)
            moved_code = moved.json()["code"]

        # Reissuing a token and activating it again from the same browser must
        # reuse the revoked device row instead of violating uq_user_device_key.
        with TestClient(app) as moved_device:
            moved_device.headers.update(
                {"X-Examify-Device-Key": "replacement-device-key-000001"}
            )
            moved_activation = moved_device.post(
                "/api/v1/activations/redeem",
                json={
                    "code": moved_code,
                    "device_key": "replacement-device-key-000001",
                },
            )
            self.assertEqual(moved_activation.status_code, 200, moved_activation.text)
            self.assertEqual(moved_activation.json()["user_id"], user_id)
            self.assertEqual(moved_activation.json()["device_id"], replacement_device_id)
            logged_in = moved_device.post(
                "/api/v1/auth/login",
                json={
                    "email": "integration-user@test.local",
                    "password": self.PASSWORD,
                    "device_key": "replacement-device-key-000001",
                },
            )
            self.assertEqual(logged_in.status_code, 200, logged_in.text)
            self.assertEqual(moved_device.get("/api/v1/exams").json()["total"], 1)

    def test_activation_token_can_bind_two_devices(self) -> None:
        with TestClient(app) as admin:
            login = admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-device-key-00000001",
                },
            )
            self.assertEqual(login.status_code, 200, login.text)
            created = admin.post(
                "/api/v1/admin/tokens",
                json={
                    "count": 1,
                    "label": "Two devices",
                    "max_devices": 2,
                    "expires_in_days": 30,
                },
            )
            self.assertEqual(created.status_code, 200, created.text)
            code = created.json()["codes"][0]

        first_key = "two-device-key-000000000001"
        second_key = "two-device-key-000000000002"
        with TestClient(app) as first:
            first_activation = first.post(
                "/api/v1/activations/redeem",
                json={"code": code, "device_key": first_key},
            )
            self.assertEqual(first_activation.status_code, 200, first_activation.text)
            owner_id = first_activation.json()["user_id"]
        with TestClient(app) as second:
            second_activation = second.post(
                "/api/v1/activations/redeem",
                json={"code": code, "device_key": second_key},
            )
            self.assertEqual(second_activation.status_code, 200, second_activation.text)
            self.assertEqual(second_activation.json()["user_id"], owner_id)
        with TestClient(app) as third:
            blocked = third.post(
                "/api/v1/activations/redeem",
                json={"code": code, "device_key": "two-device-key-000000000003"},
            )
            self.assertEqual(blocked.status_code, 409, blocked.text)
            self.assertIn("giới hạn thiết bị", blocked.json()["detail"])

        with session_scope() as session:
            token = session.scalar(
                select(models.ActivationToken).where(
                    models.ActivationToken.owner_user_id == owner_id
                )
            )
            devices = session.scalars(
                select(models.Device).where(models.Device.user_id == owner_id)
            ).all()
            self.assertIsNotNone(token)
            self.assertEqual(token.max_devices, 2)
            self.assertEqual(len(devices), 2)

    def test_revoke_token_invalidates_linked_desktop_device_and_refresh(self) -> None:
        with TestClient(app) as admin:
            login = admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-revoke-device-key-0001",
                },
            )
            self.assertEqual(login.status_code, 200, login.text)
            created = admin.post("/api/v1/admin/tokens", json={"count": 1, "label": "Desktop"})
            code = created.json()["codes"][0]
            token_id = admin.get("/api/v1/admin/tokens").json()["items"][0]["id"]

            with TestClient(app) as outdated_desktop:
                upgrade_required = outdated_desktop.post(
                    "/api/v1/desktop/activate",
                    json={
                        "code": code,
                        "device_key": "desktop-revoke-device-key-00000001",
                    },
                )
                self.assertEqual(upgrade_required.status_code, 426, upgrade_required.text)

            with TestClient(
                app, headers={"X-Examify-Desktop-Version": "0.1.2"}
            ) as desktop:
                activated = desktop.post(
                    "/api/v1/desktop/activate",
                    json={
                        "code": code,
                        "device_key": "desktop-revoke-device-key-00000001",
                        "device_name": "Examify Windows",
                        "platform": "Windows",
                    },
                )
                self.assertEqual(activated.status_code, 200, activated.text)
                desktop_tokens = self.desktop_register_and_login(
                    desktop,
                    activation=activated.json(),
                    email="desktop-revoke@test.local",
                    device_key="desktop-revoke-device-key-00000001",
                )
                wrong_machine = desktop.post(
                    "/api/v1/desktop/auth/login",
                    json={
                        "email": "desktop-revoke@test.local",
                        "password": self.PASSWORD,
                        "device_key": "different-hardware-key-00000000001",
                    },
                )
                self.assertEqual(wrong_machine.status_code, 403, wrong_machine.text)
                me = desktop.get(
                    "/api/v1/auth/me",
                    headers={"Authorization": f"Bearer {desktop_tokens['access_token']}"},
                )
                self.assertEqual(me.status_code, 200, me.text)

                revoked = admin.post(f"/api/v1/admin/tokens/{token_id}/revoke")
                self.assertEqual(revoked.status_code, 200, revoked.text)
                self.assertEqual(
                    desktop.get(
                        "/api/v1/auth/me",
                        headers={"Authorization": f"Bearer {desktop_tokens['access_token']}"},
                    ).status_code,
                    401,
                )
                refreshed = desktop.post(
                    "/api/v1/desktop/auth/refresh",
                    json={"refresh_token": desktop_tokens["refresh_token"]},
                )
                self.assertEqual(refreshed.status_code, 401, refreshed.text)

    def test_student_key_onboarding_and_role_matrix(self) -> None:
        with TestClient(app) as admin:
            admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-student-role-device-0001",
                },
            )
            created = admin.post(
                "/api/v1/admin/tokens",
                json={"count": 1, "assigned_role": "student", "exam_limit": 99},
            )
            self.assertEqual(created.status_code, 200, created.text)
            code = created.json()["codes"][0]

        device_key = "student-role-device-key-000000001"
        with TestClient(app) as student:
            student.headers.update({"X-Examify-Device-Key": device_key})
            activated = student.post(
                "/api/v1/activations/redeem",
                json={"code": code, "device_key": device_key},
            )
            self.assertEqual(activated.status_code, 200, activated.text)
            self.assertEqual(activated.json()["role"], "student")
            self.assertTrue(
                student.get("/api/v1/auth/device-status").json()["activated"]
            )
            resumed = student.post(
                "/api/v1/activations/redeem",
                json={"code": code, "device_key": device_key},
            )
            self.assertEqual(resumed.status_code, 200, resumed.text)
            self.assertTrue(resumed.json()["registration_required"])
            with TestClient(app) as wrong_device:
                rejected = wrong_device.post(
                    "/api/v1/activations/redeem",
                    json={
                        "code": code,
                        "device_key": "student-other-device-key-00000001",
                    },
                )
                self.assertEqual(rejected.status_code, 409, rejected.text)
            registered = student.post(
                "/api/v1/auth/register",
                json={
                    "display_name": "Student Matrix",
                    "email": "STUDENT-MATRIX@TEST.LOCAL",
                    "password": self.PASSWORD,
                    "password_confirmation": self.PASSWORD,
                    "role": "admin",
                },
            )
            self.assertEqual(registered.status_code, 200, registered.text)
            logged_in = student.post(
                "/api/v1/auth/login",
                json={
                    "email": "student-matrix@test.local",
                    "password": self.PASSWORD,
                    "device_key": device_key,
                },
            )
            self.assertEqual(logged_in.status_code, 200, logged_in.text)
            self.assertEqual(logged_in.json()["role"], "student")
            self.assertEqual(student.get("/api/v1/student/classrooms").status_code, 200)
            self.assertEqual(student.get("/api/v1/exams").status_code, 403)
            self.assertEqual(student.post("/api/extractions").status_code, 403)
            self.assertEqual(student.get("/api/v1/teacher/classrooms").status_code, 403)
            self.assertEqual(student.get("/api/v1/admin/dashboard").status_code, 403)
            stolen_access = student.cookies.get(auth_service.ACCESS_COOKIE)
            self.assertTrue(stolen_access)
            with TestClient(app) as wrong_machine_with_cookie:
                wrong_machine_with_cookie.cookies.set(
                    auth_service.ACCESS_COOKIE, stolen_access
                )
                wrong_machine_with_cookie.headers.update(
                    {"X-Examify-Device-Key": "stolen-session-machine-key-00001"}
                )
                rejected_session = wrong_machine_with_cookie.get("/api/v1/auth/me")
                self.assertEqual(
                    rejected_session.status_code, 401, rejected_session.text
                )
                self.assertIn(
                    "không thuộc thiết bị",
                    rejected_session.json()["detail"].lower(),
                )
            logged_out = student.post("/api/v1/auth/logout")
            self.assertEqual(logged_out.status_code, 200, logged_out.text)
            self.assertEqual(student.get("/api/v1/auth/me").status_code, 401)

        # A different browser or machine has a different device secret and must
        # never be allowed to reuse the activated account.
        with TestClient(app) as another_browser:
            another_browser.headers.update(
                {"X-Examify-Device-Key": "student-other-machine-key-000001"}
            )
            logged_in = another_browser.post(
                "/api/v1/auth/login",
                json={
                    "email": "student-matrix@test.local",
                    "password": self.PASSWORD,
                    "device_key": "student-other-machine-key-000001",
                },
            )
            self.assertEqual(logged_in.status_code, 403, logged_in.text)
            self.assertFalse(
                another_browser.get("/api/v1/auth/device-status").json()["activated"]
            )
            self.assertIn("đúng thiết bị", logged_in.json()["detail"].lower())
            with session_scope() as session:
                count = session.query(models.Device).filter_by(
                    user_id=activated.json()["user_id"]
                ).count()
                self.assertEqual(count, 1)

    def test_policies_render_legacy_markdown_and_sanitize_rich_html(self) -> None:
        with session_scope() as session:
            session.merge(
                SitePolicy(
                    key="privacy",
                    title="Legacy",
                    content="## Tiêu đề\n**Nội dung**",
                    content_format="markdown",
                )
            )
        with TestClient(app) as public:
            legacy = public.get("/api/v1/policies/privacy")
            self.assertEqual(legacy.status_code, 200, legacy.text)
            self.assertEqual(legacy.json()["content_format"], "markdown")
            self.assertIn("<h2>", legacy.json()["rendered_html"])

        with TestClient(app) as admin:
            admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-policy-device-key-00001",
                },
            )
            updated = admin.put(
                "/api/v1/policies/terms",
                json={
                    "title": "Rich policy",
                    "content_format": "html",
                    "content": '<p style="font-size: 18px">An toàn <strong>HTML</strong></p><script>alert(1)</script>',
                },
            )
            self.assertEqual(updated.status_code, 200, updated.text)
            payload = updated.json()
            self.assertEqual(payload["content_format"], "html")
            self.assertIn("font-size: 18px", payload["rendered_html"])
            self.assertNotIn("script", payload["rendered_html"].lower())

    def test_guides_admin_lifecycle_public_visibility_and_xss_sanitizing(self) -> None:
        guide_id = ""
        with TestClient(app) as anonymous:
            forbidden = anonymous.get("/api/v1/admin/guides")
            self.assertEqual(forbidden.status_code, 401, forbidden.text)
            self.assertEqual(anonymous.get("/api/v1/guides").status_code, 401)

        with TestClient(app) as admin:
            login = admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-guide-device-key-00001",
                },
            )
            self.assertEqual(login.status_code, 200, login.text)
            category = admin.post(
                "/api/v1/admin/guide-categories",
                json={"name": "Bắt đầu", "slug": "bat-dau", "sort_order": 1},
            )
            self.assertEqual(category.status_code, 201, category.text)
            created = admin.post(
                "/api/v1/admin/guides",
                json={
                    "title": "Hướng dẫn đầu tiên",
                    "slug": "huong-dan-dau-tien",
                    "summary": "Mô tả có thể tìm kiếm",
                    "category_id": category.json()["id"],
                    "content": {"type": "doc", "content": [{"type": "paragraph"}]},
                    "rendered_html": (
                        '<h2>Khởi động</h2><p style="color:#123456">An toàn</p>'
                        '<a href="javascript:alert(1)">xấu</a><script>alert(2)</script>'
                    ),
                    "status": "DRAFT",
                    "keywords": ["cài đặt"],
                },
            )
            self.assertEqual(created.status_code, 201, created.text)
            guide_id = created.json()["id"]
            self.assertNotIn("script", created.json()["rendered_html"].lower())
            self.assertNotIn("javascript:", created.json()["rendered_html"].lower())

        with TestClient(app) as public:
            self.assertEqual(public.get("/api/v1/guides").status_code, 401)

        with TestClient(app) as admin:
            admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-guide-device-key-00001",
                },
            )
            published = admin.post(f"/api/v1/admin/guides/{guide_id}/publish")
            self.assertEqual(published.status_code, 200, published.text)

        with TestClient(app) as admin:
            admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-guide-device-key-00001",
                },
            )
            listing = admin.get("/api/v1/guides", params={"q": "cài đặt"})
            self.assertEqual(listing.status_code, 200, listing.text)
            self.assertEqual(listing.json()["total"], 1)
            searched = admin.get("/api/v1/guides/search", params={"q": "cài đặt"})
            self.assertEqual(searched.status_code, 200, searched.text)
            self.assertEqual(searched.json()["total"], 1)
            detail = admin.get("/api/v1/guides/huong-dan-dau-tien")
            self.assertEqual(detail.status_code, 200, detail.text)
            self.assertIn("<h2>Khởi động</h2>", detail.json()["rendered_html"])

        with TestClient(app) as admin:
            admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-guide-device-key-00001",
                },
            )
            hidden = admin.post(f"/api/v1/admin/guides/{guide_id}/unpublish")
            self.assertEqual(hidden.json()["status"], "HIDDEN")
            removed = admin.delete(f"/api/v1/admin/guides/{guide_id}")
            self.assertEqual(removed.status_code, 200, removed.text)

    def test_exam_quota_and_permanent_admin_deletion(self) -> None:
        label = "Quota deletion test"
        with TestClient(app) as admin:
            login = admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-quota-device-key-00001",
                },
            )
            self.assertEqual(login.status_code, 200, login.text)
            created = admin.post(
                "/api/v1/admin/tokens",
                json={
                    "count": 1,
                    "label": label,
                    "exam_limit": 1,
                    "expires_in_days": 30,
                },
            )
            self.assertEqual(created.status_code, 200, created.text)
            code = created.json()["codes"][0]
            token = next(
                item
                for item in admin.get("/api/v1/admin/tokens").json()["items"]
                if item["label"] == label
            )
            self.assertEqual(token["exam_limit"], 1)

        with TestClient(app) as user:
            activated = user.post(
                "/api/v1/activations/redeem",
                json={
                    "code": code,
                    "device_key": "quota-user-device-key-00001",
                    "device_name": "Quota test",
                },
            )
            self.assertEqual(activated.status_code, 200, activated.text)
            user_id = activated.json()["user_id"]
            self.register_and_login(
                client=user,
                email="quota-user@test.local",
                device_key="quota-user-device-key-00001",
            )

        payload = {
            "schema_version": 2,
            "exam_type": "reading",
            "questions": [],
            "stimuli": [],
            "audios": [],
        }
        first_exam_id = persist_final_exam(
            payload,
            job_id=None,
            owner_user_id=user_id,
            title="Quota exam 1",
        )
        with self.assertRaises(HTTPException) as blocked:
            persist_final_exam(
                payload,
                job_id=None,
                owner_user_id=user_id,
                title="Quota exam 2",
            )
        self.assertEqual(blocked.exception.status_code, 403)
        with session_scope() as session:
            session.get(models.Exam, first_exam_id).deleted_at = models.utcnow()
        with self.assertRaises(HTTPException):
            persist_final_exam(
                payload,
                job_id=None,
                owner_user_id=user_id,
                title="Quota exam after deletion",
            )

        with TestClient(app) as admin:
            admin.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-quota-device-key-00001",
                },
            )
            increased = admin.patch(
                f"/api/v1/admin/users/{user_id}",
                json={"exam_limit": 2},
            )
            self.assertEqual(increased.status_code, 200, increased.text)
            second_exam_id = persist_final_exam(
                payload,
                job_id=None,
                owner_user_id=user_id,
                title="Quota exam 2",
            )
            self.assertIsNotNone(second_exam_id)

            deleted_token = admin.delete(f"/api/v1/admin/tokens/{token['id']}")
            self.assertEqual(deleted_token.status_code, 200, deleted_token.text)
            token_ids = {
                item["id"]
                for item in admin.get("/api/v1/admin/tokens").json()["items"]
            }
            self.assertNotIn(token["id"], token_ids)

            deleted_user = admin.delete(f"/api/v1/admin/users/{user_id}")
            self.assertEqual(deleted_user.status_code, 200, deleted_user.text)
            user_ids = {
                item["id"]
                for item in admin.get("/api/v1/admin/users").json()["items"]
            }
            self.assertNotIn(user_id, user_ids)
            with session_scope() as session:
                self.assertIsNone(session.get(models.User, user_id))
                self.assertIsNone(session.get(models.Exam, first_exam_id))
                self.assertIsNone(session.get(models.Exam, second_exam_id))

    def test_current_identity_optional_and_asset_access_without_device_key(self) -> None:
        """HTML media gets a signed internal MinIO redirect without a device header."""
        with session_scope() as session:
            owner_id = session.scalar(
                select(User.id).where(User.email == config.settings.admin_email)
            )
        exam_id = persist_final_exam(
            {
                "schema_version": 2,
                "exam_type": "reading",
                "questions": [],
                "stimuli": [],
                "audios": [],
            },
            job_id=None,
            owner_user_id=owner_id,
            title="Public Asset Test Exam",
        )
        with session_scope() as session:
            session.add(
                models.Asset(
                    exam_id=exam_id,
                    kind="stimulus",
                    bucket="examify-assets",
                    object_key="test-key",
                    filename="sample.webp",
                    content_type="image/webp",
                    size=10,
                )
            )

        with TestClient(app) as client:
            # Login sets smart_exam_access cookie
            client.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": "admin-asset-test-device-key-0001",
                },
            )

            # Standard HTML img tag request HAS smart_exam_access cookie BUT NO X-Examify-Device-Key header
            self.assertNotIn("X-Examify-Device-Key", client.headers)
            fake_storage = Mock()
            fake_storage.presigned_internal_redirect.return_value = (
                "/_protected_minio/examify-assets/test-key"
                "?X-Amz-Algorithm=AWS4-HMAC-SHA256&X-Amz-Signature=signed"
            )
            with patch.object(platform_api, "storage", fake_storage), patch.object(
                platform_api,
                "settings",
                SimpleNamespace(minio_accel_redirect_prefix="/_protected_minio"),
            ):
                res = client.get(f"/api/v1/exams/{exam_id}/assets/sample.webp")

            self.assertEqual(res.status_code, 200, res.text)
            self.assertIn("X-Amz-Signature=signed", res.headers["X-Accel-Redirect"])
            fake_storage.presigned_internal_redirect.assert_called_once_with(
                "examify-assets",
                "test-key",
                "/_protected_minio",
                method="GET",
            )

    def test_public_mini_test_share_and_submission(self) -> None:
        """Verify teacher can generate public share link and students can take the exam anonymously."""
        with session_scope() as session:
            admin_user = session.scalar(
                select(User).where(User.email == config.settings.admin_email)
            )
            owner_id = admin_user.id
        exam_id = persist_final_exam(
            {
                "schema_version": 2,
                "exam_type": "reading",
                "category": "Mini Test",
                "questions": [
                    {
                        "number": 1,
                        "part": "Part 5",
                        "text": "Choose A",
                        "options": {"A": "A", "B": "B", "C": "C", "D": "D"},
                        "correct": "A",
                    }
                ],
                "stimuli": [],
                "audios": [],
            },
            job_id=None,
            owner_user_id=owner_id,
            title="Mini Test 1",
        )

        with TestClient(app) as client:
            device_key = "admin-mini-test-device-key-0001"
            login_res = client.post(
                "/api/v1/auth/login",
                json={
                    "email": "admin@test.local",
                    "password": "test-admin-password",
                    "device_key": device_key,
                },
            )
            self.assertEqual(login_res.status_code, 200, login_res.text)
            client.headers["X-Examify-Device-Key"] = device_key

            # Teacher creates public share link
            res_share = client.post(f"/api/v1/exams/{exam_id}/public-share")
            self.assertEqual(res_share.status_code, 200, res_share.text)
            share_data = res_share.json()
            share_code = share_data["share_code"]
            self.assertTrue(share_code.startswith("mini-"))

            # Anonymous student fetches test details (no auth cookies or device keys required)
            anon_client = TestClient(app)
            res_details = anon_client.get(f"/api/v1/public-tests/{share_code}")
            self.assertEqual(res_details.status_code, 200, res_details.text)
            self.assertEqual(res_details.json()["title"], "Mini Test 1")
            self.assertNotIn("correct", res_details.json()["exam"]["questions"][0])

            # Student inputs name and starts test
            res_start = anon_client.post(
                f"/api/v1/public-tests/{share_code}/start",
                json={"student_name": "Nguyễn Văn A", "phone": "0987654321"},
            )
            self.assertEqual(res_start.status_code, 200, res_start.text)
            sub_id = res_start.json()["submission_id"]
            submission_token = res_start.json()["submission_token"]

            rejected = anon_client.post(
                f"/api/v1/public-tests/submissions/{sub_id}/submit",
                json={"submission_token": "x" * 32, "answers": {"1": "A"}},
            )
            self.assertEqual(rejected.status_code, 401, rejected.text)

            # Student submits test answers
            res_submit = anon_client.post(
                f"/api/v1/public-tests/submissions/{sub_id}/submit",
                json={
                    "submission_token": submission_token,
                    "answers": {"1": "A"},
                    "time_spent_seconds": 300,
                },
            )
            self.assertEqual(res_submit.status_code, 200, res_submit.text)
            sub_data = res_submit.json()
            self.assertEqual(sub_data["total_correct"], 1)
            self.assertEqual(sub_data["part_breakdown"]["Part 5"]["correct"], 1)

            # Teacher checks public submissions list
            res_list = client.get(f"/api/v1/exams/{exam_id}/public-submissions")
            self.assertEqual(res_list.status_code, 200, res_list.text)
            items = res_list.json()["items"]
            self.assertEqual(len(items), 1)
            self.assertEqual(items[0]["student_name"], "Nguyễn Văn A")
            self.assertEqual(items[0]["phone"], "0987654321")


if __name__ == "__main__":
    unittest.main()
